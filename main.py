import requests
import pandas as pd
import argparse
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

def k_parameter():
    parser = argparse.ArgumentParser(description="Process keys and colored flag")
 
    # Add the -k/--keys parameter that accepts multiple string arguments
    parser.add_argument("-k", "--keys", nargs='+', help="List of keys")

    # Add the -c/--colored flag with a default value of True
    parser.add_argument("-c", "--colored", action="store_false", default=True, help="Enable or disable colored output")

    # Parse the command-line arguments
    args = parser.parse_args()
    
    return args.keys, args.colored
    
def assign_color_code(date):
    date = datetime.strptime(date, '%Y-%m-%d')  # Convert date to a datetime object
    days_difference = (datetime.now() - date).days

    if days_difference <= 90:  # If not older than 3 months
        return '007500'  # Green
    elif days_difference <= 365:  # If not older than 12 months
        return 'FFA500'  # Orange
    else:
        return 'b30000'  # Red
    
def get_column():
    #Arranges columns regarding inputs
    columns=['rnr']
    if keys!=None:
        for key in keys:
            columns.append(key)
    return columns

def labelIds_control(columns):
    #Controls and returns index of labelIds
    indexes=[]
    for column in columns:
        if column=='labelIds':
            filter = response_df['labelIds']!=''
            indexes = response_df[filter].index

    return indexes

def column_toexcel(ws):
    #Enters columns values to Excel File
    for col_num, column_title in enumerate(response_df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

def rows_toexcel(ws):
    #Enters rows values to Excel File
    for row_num, (_, row_data) in enumerate(response_df.iterrows(), 2):
        for col_num, cell_value in enumerate(row_data, 1):
            if response_df.columns[col_num - 1] == 'labelIds':
                color_code = row_data['labelIds']  
                #Color code control and colourizing texts of row related to labelIds value
                if color_code!='':
                    cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                    cell.font = Font(color=color_code)  
            
            ws.cell(row=row_num, column=col_num, value=cell_value)

def color_cell(c,ws):
    #c value control and Colorize backgrounds of cells 
    if c==True:
        for col_num, column_name in enumerate(response_df.columns, 1):
            for row_num, cell_value in enumerate(response_df[column_name], 2):
                date_hu = date_df.iloc[row_num - 2]
                color_code=assign_color_code(date_hu)

                fill = PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.fill = fill


keys,c=k_parameter()

base_url = "http://127.0.0.1:8000"
csv_file_path = "vehicles.csv"

with open(csv_file_path, "rb") as csv_file:
    response = requests.post(base_url, files={"file": ("vehicles.csv", csv_file)})

response_df=pd.DataFrame(response.json())
response_df=response_df.sort_values(by='gruppe')

columns=get_column()
indexes=labelIds_control(columns)


date_df=response_df['hu']
response_df=response_df[columns]
wb = Workbook()
ws = wb.active

column_toexcel(ws)
rows_toexcel(ws)
color_cell(c,ws)

current_date = datetime.now().strftime('%Y-%m-%d')
file_name = f'data_{current_date}.xlsx'
wb.save(file_name)
